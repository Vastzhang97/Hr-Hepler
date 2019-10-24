from selenium import webdriver
import time
from PIL import Image
from selenium.common.exceptions import UnexpectedAlertPresentException, NoAlertPresentException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from util import Util
import re
import util.StringUtil as StringUil
import os, sys
from util.BaiduApiUtil import BaiDuApi
from openpyxl import load_workbook  # pip install openpyxl

strUtil = StringUil.StringUtil()
file = "C://Users/62526/Desktop/Test/职称信息1.xlsx"
wb = load_workbook(file)
wb.guess_types = True
ws = wb.active
name_column = 2
id_num_column = 3
column4 = 4  # 在职单位名称
column5 = 5  # 学历
column6 = 6  # #注册证书信息
column7 = 7  # 注册类别
column8 = 8  # 注册号
column9 = 9  # 注册单位
column10 = 10  # 发证机关
column11 = 11  # 签发日期
column12 = 12  # 有效期
column13 = 13  # #职称证书信息
column14 = 14  # 证书名称
column15 = 15  # 职称等级
column16 = 16  # 发证机关
column17 = 17  # 发证日期
column18 = 18  # 有效期
column19 = 19  # 安全生产考核合格证
column20 = 20  # 证书名称
column21 = 21  # 证书编号
column22 = 22  # 发证机关
column23 = 23  # 发证日期
column24 = 24  # 有效期
column25 = 25  # 岗位证书信息
column26 = 26  # 证书名称
column27 = 27  # 证书编号
column28 = 28  # 发证机关
column29 = 29  # 发证日期
column30 = 30  # 有效期
column31 = 31  # 是否爬取
column32 = 32  # url
current_max_row = ws.max_row
current_row = 2
data_first_row = 2


def search(name, id_num):
    name_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtPersonName']")
    name_input.clear()
    name_input.send_keys(name)
    num_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtIdnum']")
    num_input.clear()
    num_input.send_keys(id_num)
    search_btn = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_btnSearch']")
    verification_code_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtCheckCode']")
    verification_code = driver.find_element_by_xpath("//a/img[@id='CheckCodeImage']")
    driver.save_screenshot("screenshot.png")
    location = verification_code.location
    size = verification_code.size
    x = location["x"]
    y = location["y"]
    width = size["width"]
    height = size["height"]
    range = (int(x + 160), int(y + 90), int(x + width + 165),
             int(y + height + 85))
    # im = im.crop((left, top, right, bottom))
    im = Image.open("screenshot.png")
    im = im.crop(range)
    im = im.convert("RGB")
    im.save("code.png")
    api = BaiDuApi()
    code = api.read_image("code.png")
    # code = "test"
    os.remove("screenshot.png")
    os.remove("code.png")
    time.sleep(0.5)
    verification_code_input.send_keys(code)
    search_btn.click()
    driver.switch_to.alert.accept()
    search(name, id_num)


def get_result(html_str):
    start_str1 = '<div class="spec-title">职称证书信息</div>'
    end_str1 = '<!--列表最多显示5条，5条以上的显示更多按钮-->'
    pos1 = html_str.index(start_str1)
    pos2 = html_str.index(end_str1, pos1)
    html_str = html_str[pos1:pos2]
    if html_str.find("暂无信息") > 0:
        return ""
    start_str2 = "<td>"
    end_str2 = "</td>"
    num = html_str.count(start_str2)
    pos3 = 0
    pos4 = 0
    for i in range(num):
        pos3 = html_str.find(start_str2, pos3) + 4
        pos4 = html_str.find(end_str2, pos3)
        if i % 5 is 0:
            certificate_name = html_str[pos3:pos4]
        if i % 5 is 1:
            level = html_str[pos3:pos4]
        if i % 5 is 2:
            licence_issuing = html_str[pos3:pos4]
        if i % 5 is 3:
            date = html_str[pos3:pos4]
        else:
            validity = html_str[pos3:pos4]


# 注册证书信息
def get_registration_information(information, current_row, insert_row_num):
    str1 = '<div class="spec-title">职称证书信息</div>'
    str2 = '<!--列表最多显示5条，5条以上的显示更多按钮-->'
    str3 = "<td>"
    str4 = "</td>"
    print("注册证书信息")
    data_num = 6
    if information.find("暂无信息") > 0:
        driver.close()
        driver.switch_to.window(index_handle)
        print("注册证书信息暂无信息")
        return ""
    information = strUtil.cut_str_between_two_str(information, str1, str2)
    str_list = strUtil.get_all_str_beween_two_str(information, str3, str4)
    for index, item in enumerate(str_list):
        if index is not 0 and index % data_num is 0:
            current_row += 1
            ws.insert_rows(current_row)
            insert_row_num += 1
        if index % data_num is 0:
            print("--------------------------")
            print("注册类别 " + item)
            ws.cell(current_row, column5).value = item
        elif index % data_num is 1:
            print("注册号 " + item)
            ws.cell(current_row, column15).value = item
        elif index % data_num is 2:
            print("注册单位 " + item)
            ws.cell(current_row, column16).value = item
        elif index % data_num is 3:
            print("发证机关 " + item)
            ws.cell(current_row, column17).value = item
        elif index % data_num is 4:
            print("签发日期 " + item)
            ws.cell(current_row, column17).value = item
        else:
            print("有效期 " + item)
            print("--------------------------")
            ws.cell(current_row, column18).value = item
    return information


# 职称证书信息
def get_professional_title_certificate_information():
    pass


# 安全生产考核合格证
def get_safety_production_assessment_certificate():
    pass


# 岗位证书信息
def get_post_certificate_information():
    pass


util = Util.Util()
driver = webdriver.Chrome()
url = "http://113.108.219.40/Dop/Open/PersonList.aspx"
driver.implicitly_wait(10)
driver.maximize_window()
driver.get(url)
time.sleep(2)

for i in range(1, current_max_row):
    insert_row_num = 0
    if i is 1:
        continue
    has_crawler_flag = ws.cell(current_row, column31).value
    name = ws.cell(current_row, name_column).value
    id_num = ws.cell(current_row, id_num_column).value
    if has_crawler_flag is None:
        print("--------------------------")
        print("name " + name)
        print("id_num " + id_num)
        print("--------------------------")
        if name is not None and id_num is not None:
            try:
                search(name, id_num)
            except NoAlertPresentException as e:
                pass

            result_list = driver.find_elements_by_xpath("//table[@class='data-list']/tbody/tr/td/a")
            index_handle = driver.current_window_handle
            for index, item in enumerate(result_list):
                time.sleep(1)
                item.click()
                util.switch_to_new_window(driver, index_handle)
                time.sleep(2)
                url = driver.current_url
                print("url " + url)
                ws.cell(current_row, column32).value = url
                try:
                    company_div = driver.find_element_by_xpath(
                        "//div[@class='spec-item-4'][4]/div[@class='spec-item']/h5/a")
                    company = company_div.get_attribute("innerHTML")
                    ws.cell(current_row, column4).value = company
                except NoSuchElementException as e:
                    print("无在职单位")
                try:
                    educational_div = driver.find_element_by_xpath(
                        "//div[@class='spec-item-4'][3]/div[@class='spec-item']/h5")
                    educational = educational_div.get_attribute("innerHTML")
                    ws.cell(current_row, column5).value = educational
                except NoSuchElementException as e:
                    print("无学历")
                information_div = driver.find_elements_by_xpath("//div[@class='spec-display']")
                information = ""
                for index3, item3 in enumerate(information_div):
                    information += item3.get_attribute("outerHTML")

                registration_information = get_registration_information \
                    (information, current_row, insert_row_num)

                driver.close()
                driver.switch_to.window(index_handle)
    # data_first_row = data_first_row + insert_row_num
    ws.cell(current_row, column31).value = name
    wb.save(file)
    data_first_row += 1
    current_row += 1

driver.quit()
