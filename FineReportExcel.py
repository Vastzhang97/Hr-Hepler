from openpyxl import load_workbook  # pip install openpyxl

file = "C:/Users/developer/Desktop/file.xlsx"
wb = load_workbook(file)
wb.guess_types = True
ws = wb.active
current_max_row = ws.max_row
markdown_str = "功能|站点|birt报表名称|finereport报表路径|备注|迁移人员"+"\n"
markdown_str += ":-:|:-:|:-:|:-:|:-:|:-:" + "\n"
for i in range(current_max_row):
    if i is 0 or i is 1:
        continue
    if ws.cell(i, 1).value is not None:
        markdown_str += ws.cell(i, 1).value.replace("\n", " ").replace("\r", " ")
    else:
        markdown_str += "-"
    markdown_str += "|"
    if ws.cell(i, 2).value is not None:
        markdown_str += ws.cell(i, 2).value.replace("\n", " ").replace("\r", " ")
    else:
        markdown_str += "-"
    markdown_str += "|"
    if ws.cell(i, 3).value is not None:
        markdown_str += ws.cell(i, 3).value.replace("\n", " ").replace("\r", " ")
    else:
        markdown_str += "-"
    markdown_str += "|"
    if ws.cell(i, 4).value is not None:
        markdown_str += ws.cell(i, 4).value.replace("\n", " ").replace("\r", " ")
    else:
        markdown_str += "-"
    markdown_str += "|"
    if ws.cell(i, 5).value is not None:
        markdown_str += ws.cell(i, 5).value.replace("\n", " ").replace("\r", " ")
    else:
        markdown_str += "-"
    markdown_str += "|张瀚" + "\n"
print(markdown_str)
