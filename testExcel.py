from openpyxl import Workbook  # pip install openpyxl
from openpyxl import load_workbook

file = "C://Users/62526/Desktop/Test/职称信息1.xlsx"
wb = load_workbook(file)
wb.guess_types = True
ws = wb.active
name_column = 2
id_num_column = 3
company_column = 4
certificate_name_column = 5
level_column = 6
licence_issuing_column = 7
date_column = 8
validity_column = 9
has_crawling_column = 10

for i in range(1, ws.max_row):
    if i is 1:
        continue
    value = ws.cell(i, has_crawling_column).value
    if value is not None:
        print(value)
        name = ws.cell(i, name_column).value
        id_num = ws.cell(i, id_num_column).value
        ws.cell(i, has_crawling_column).value = "是"
# ws.cell(row=1, column=1).value = "test"
# ws.insert_rows(2)
# print(ws.cell(row=1, column=1).value)
# print(ws.title)

# wb.save(file)
