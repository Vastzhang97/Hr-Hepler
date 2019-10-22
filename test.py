from selenium import webdriver
import time
from PIL import Image
from selenium.common.exceptions import UnexpectedAlertPresentException, NoAlertPresentException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from util import Util
import re
import os, sys
from util.BaiduApiUtil import BaiDuApi
from openpyxl import load_workbook  # pip install openpyxl


def search(name, id_num):
    name_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtPersonName']")
    name_input.clear()
    name_input.send_keys(name)
    num_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtIdnum']")
    num_input.clear()
    num_input.send_keys(id_num)
    search_btn = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_btnSearch']")
    verification_code_input = driver.find_element_by_xpath("//input[@id='ctl00_ContentPlaceHolder1_txtCheckCode']")
    verification_code = driver.find_element_by_xpath("//a/img[@id='CheckCodeImage']")
    driver.save_screenshot("screenshot.png")
    location = verification_code.location
    size = verification_code.size
    x = location["x"]
    y = location["y"]
    width = size["width"]
    height = size["height"]
    range = (int(x + 160), int(y + 90), int(x + width + 165),
             int(y + height + 85))
    # im = im.crop((left, top, right, bottom))
    im = Image.open("screenshot.png")
    im = im.crop(range)
    im = im.convert("RGB")
    im.save("code.png")
    api = BaiDuApi()
    code = api.read_image("code.png")
    # code = "test"
    os.remove("screenshot.png")
    os.remove("code.png")
    time.sleep(0.5)
    verification_code_input.send_keys(code)
    search_btn.click()
    driver.switch_to.alert.accept()
    search(name, id_num)


def get_result(html_str):
    start_str1 = '<div class="spec-title">职称证书信息</div>'
    end_str1 = '<!--列表最多显示5条，5条以上的显示更多按钮-->'
    pos1 = html_str.index(start_str1)
    pos2 = html_str.index(end_str1, pos1)
    html_str = html_str[pos1:pos2]
    if html_str.find("暂无信息") > 0:
        return ""
    start_str2 = "<td>"
    end_str2 = "</td>"
    num = html_str.count(start_str2)
    pos3 = 0
    pos4 = 0
    for i in range(num):
        pos3 = html_str.find(start_str2, pos3) + 4
        pos4 = html_str.find(end_str2, pos3)
        if i % 5 is 0:
            certificate_name = html_str[pos3:pos4]
        if i % 5 is 1:
            level = html_str[pos3:pos4]
        if i % 5 is 2:
            licence_issuing = html_str[pos3:pos4]
        if i % 5 is 3:
            date = html_str[pos3:pos4]
        else:
            validity = html_str[pos3:pos4]


file = "C://Users/62526/Desktop/Test/职称信息1.xlsx"
wb = load_workbook(file)
wb.guess_types = True
ws = wb.active
name_column = 2
id_num_column = 3
company_column = 4
certificate_name_column = 5
level_column = 6
licence_issuing_column = 7
date_column = 8
validity_column = 9
has_crawling_column = 10

util = Util.Util()
driver = webdriver.Chrome()
url = "http://113.108.219.40/Dop/Open/PersonList.aspx"
driver.implicitly_wait(10)
driver.maximize_window()
driver.get(url)
time.sleep(2)

current_max_row = ws.max_row
current_row = 2
data_first_row = 2
for i in range(1, current_max_row):
    if i is 1:
        continue
    value = ws.cell(current_row, has_crawling_column).value
    if value is None:
        name = ws.cell(current_row, name_column).value
        id_num = ws.cell(current_row, id_num_column).value
        print("--------------------------")
        print("name " + name)
        print("id_num " + id_num)
        print("--------------------------")
        if name is not None and id_num is not None:
            try:
                search(name, id_num)
            except NoAlertPresentException as e:
                pass

            result_list = driver.find_elements_by_xpath("//table[@class='data-list']/tbody/tr/td/a")
            index_handle = driver.current_window_handle
            for index, item in enumerate(result_list):
                time.sleep(1)
                item.click()
                util.switch_to_new_window(driver, index_handle)
                company_div = driver.find_element_by_xpath(
                    "//div[@class='spec-item-4'][4]/div[@class='spec-item']/h5/a")
                company = company_div.get_attribute("innerHTML")
                ws.cell(current_row, company_column).value = company
                information_div = driver.find_elements_by_xpath("//div[@class='spec-display']")
                information = ""
                for index3, item3 in enumerate(information_div):
                    information += item3.get_attribute("outerHTML")
                start_str2 = "<td>"
                end_str2 = "</td>"

                start_str1 = '<div class="spec-title">职称证书信息</div>'
                end_str1 = '<!--列表最多显示5条，5条以上的显示更多按钮-->'
                pos1 = information.index(start_str1)
                pos2 = information.index(end_str1, pos1)
                information = information[pos1:pos2]
                if information.find("暂无信息") > 0:
                    driver.close()
                    driver.switch_to.window(index_handle)
                    print("暂无信息")
                    break
                num = information.count(start_str2)
                pos3 = 0
                pos4 = 0
                insert_row_num = 0
                for j in range(num):
                    if j is not 0 and j % 5 is 0:
                        current_row += 1
                        ws.insert_rows(current_row)
                        insert_row_num += 1
                    pos3 = information.find(start_str2, pos3) + 4
                    pos4 = information.find(end_str2, pos3)
                    if j % 5 is 0:
                        certificate_name = information[pos3:pos4]
                        print("--------------------------")
                        print("certificate_name " + certificate_name)
                        ws.cell(current_row, certificate_name_column).value = certificate_name
                    elif j % 5 is 1:
                        level = information[pos3:pos4]
                        print("level " + level)
                        ws.cell(current_row, level_column).value = level
                    elif j % 5 is 2:
                        licence_issuing = information[pos3:pos4]
                        print("licence_issuing " + licence_issuing)
                        ws.cell(current_row, licence_issuing_column).value = licence_issuing
                    elif j % 5 is 3:
                        date = information[pos3:pos4]
                        print("date " + date)
                        ws.cell(current_row, date_column).value = date
                    else:
                        validity = information[pos3:pos4]
                        print("validity " + validity)
                        print("--------------------------")
                        ws.cell(current_row, validity_column).value = validity
                driver.close()
                driver.switch_to.window(index_handle)
    data_first_row = data_first_row + insert_row_num
    ws.cell(data_first_row, has_crawling_column).value = "是"
    wb.save(file)
    data_first_row += 1
    current_row += 1

driver.quit()
